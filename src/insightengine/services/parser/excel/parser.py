from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

try:  # pragma: no cover - 测试环境缺失时提示
    import xlrd
except ModuleNotFoundError:  # pragma: no cover
    xlrd = None  # type: ignore[assignment]


from ..base import Parser, ParserError
from ..registry import register_parser
from ..text_utils import normalize_whitespace
from ..types import ParseItem, ParseResult


_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLS_CONTENT_TYPE = "application/vnd.ms-excel"


@dataclass
class _ExcelSource:
    extension: str
    path: str | None
    buffer: io.BytesIO | None

    def openpyxl_input(self) -> str | io.BytesIO:
        if self.buffer is not None:
            self.buffer.seek(0)
            return self.buffer
        if self.path is None:
            raise ParserError("missing data buffer for openpyxl input")
        return self.path

    def xlrd_kwargs(self) -> dict[str, Any]:
        if self.buffer is not None:
            self.buffer.seek(0)
            return {"file_contents": self.buffer.getvalue()}
        if self.path is None:
            raise ParserError("missing data buffer for xlrd input")
        return {"filename": self.path}


class ExcelParser(Parser):
    """解析 Excel `.xlsx` 与 `.xls` 文件。"""

    name = "excel"

    def parse(self, source: Any, **opts) -> ParseResult:
        excel_source = _coerce_excel_source(source, **opts)

        if excel_source.extension == ".xlsx":
            workbook = load_workbook(
                filename=excel_source.openpyxl_input(),
                read_only=True,
                data_only=True,
            )
            try:
                items, sheet_names, row_stats = _extract_from_openpyxl(workbook)
            finally:
                workbook.close()
            content_type = _XLSX_CONTENT_TYPE
        elif excel_source.extension == ".xls":
            if xlrd is None:
                raise ParserError(
                    "xlrd is required to parse legacy .xls files; install xlrd>=2.0"
                )
            workbook = xlrd.open_workbook(**excel_source.xlrd_kwargs())  # type: ignore[misc]
            items, sheet_names, row_stats = _extract_from_xlrd(workbook)
            content_type = _XLS_CONTENT_TYPE
        else:  # pragma: no cover - 逻辑上不会触发
            raise ParserError(f"unsupported Excel extension: {excel_source.extension}")

        metadata_payload = {
            "parser": self.name,
            "content_type": content_type,
            "sheet_names": sheet_names,
            "sheet_count": len(sheet_names),
            "row_stats": row_stats,
        }
        extra_metadata = opts.get("metadata", {})
        metadata_payload.update(extra_metadata)

        return ParseResult(
            source=excel_source.path,
            items=items,
            metadata=metadata_payload,
        )


def _coerce_excel_source(source: Any, **opts: Any) -> _ExcelSource:
    extension = _determine_extension(source, opts)
    buffer: io.BytesIO | None = None
    path: str | None = None

    if isinstance(source, Path):
        path = str(source)
    elif isinstance(source, str):
        candidate = Path(source)
        if candidate.exists():
            path = str(candidate)
        else:
            raise ParserError("string source for Excel must be an existing file path")
    elif isinstance(source, (bytes, bytearray)):
        buffer = io.BytesIO(bytes(source))
    elif hasattr(source, "read"):
        data = source.read()
        if isinstance(data, str):
            data = data.encode("utf-8")
        if not isinstance(data, (bytes, bytearray)):
            raise ParserError(
                "file-like object for Excel must return bytes or str from read()"
            )
        buffer = io.BytesIO(bytes(data))
        path = getattr(source, "name", None)
        if path:
            extension = extension or Path(path).suffix.lower()
    else:
        raise ParserError(f"unsupported Excel source type: {type(source)!r}")

    if not extension:
        raise ParserError(
            "unable to determine Excel extension; provide extension='.xlsx' or '.xls'"
        )

    normalized_ext = extension.lower()
    if normalized_ext not in {".xlsx", ".xls"}:
        raise ParserError(f"unsupported Excel extension: {normalized_ext}")

    return _ExcelSource(extension=normalized_ext, path=path, buffer=buffer)


def _determine_extension(source: Any, opts: dict[str, Any]) -> str | None:
    ext = opts.get("extension")
    if isinstance(ext, str) and ext:
        return ext if ext.startswith(".") else f".{ext}"

    content_type = opts.get("content_type")
    if isinstance(content_type, str):
        content_type = content_type.lower()
        if content_type == _XLSX_CONTENT_TYPE:
            return ".xlsx"
        if content_type == _XLS_CONTENT_TYPE:
            return ".xls"

    if isinstance(source, (Path, str)):
        path = Path(source)
        if path.suffix:
            return path.suffix.lower()

    return None


def _build_row_item(
    sheet: str,
    row_index: int,
    values: Iterable[Any],
    position: int,
) -> ParseItem | None:
    cell_texts = []
    for value in values:
        if value is None:
            continue
        text = normalize_whitespace(str(value))
        if text:
            cell_texts.append(text)

    if not cell_texts:
        return None

    row_text = normalize_whitespace(" | ".join(cell_texts))
    if not row_text:
        return None

    return ParseItem(
        id=f"{sheet}-row-{row_index}",
        text=row_text,
        length=len(row_text),
        position=position,
        metadata={"sheet": sheet, "row": row_index, "columns": len(cell_texts)},
    )


def _extract_from_openpyxl(
    workbook,
) -> tuple[list[ParseItem], list[str], dict[str, int]]:
    items: list[ParseItem] = []
    sheet_names: list[str] = []
    row_stats: dict[str, int] = {}
    position = 1

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        sheet_names.append(sheet_name)
        processed_rows = 0

        for row_index, row in enumerate(
            worksheet.iter_rows(values_only=True),
            start=1,
        ):
            item = _build_row_item(sheet_name, row_index, row, position)
            if item is None:
                continue
            items.append(item)
            position += 1
            processed_rows += 1

        row_stats[sheet_name] = processed_rows

    return items, sheet_names, row_stats


def _extract_from_xlrd(workbook) -> tuple[list[ParseItem], list[str], dict[str, int]]:
    items: list[ParseItem] = []
    sheet_names: list[str] = []
    row_stats: dict[str, int] = {}
    position = 1

    for sheet in workbook.sheets():
        sheet_name = sheet.name
        sheet_names.append(sheet_name)
        processed_rows = 0

        for row_index in range(1, sheet.nrows + 1):
            values = sheet.row_values(row_index - 1)
            item = _build_row_item(sheet_name, row_index, values, position)
            if item is None:
                continue
            items.append(item)
            position += 1
            processed_rows += 1

        row_stats[sheet_name] = processed_rows

    return items, sheet_names, row_stats


register_parser(ExcelParser.name, ExcelParser)

__all__ = ["ExcelParser"]
